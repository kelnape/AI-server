import customtkinter as ctk
import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
import threading
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

def cista_cena(cena_str):
    cisla = re.sub(r'[^\d]', '', cena_str)
    return int(cisla) if cisla else 99999999 

class BazosScraperApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Vyhledávač Bazoš PRO - Hlídací pes")
        self.geometry("550x700")
        
        self.soubor_historie = "historie_odkazu.txt"
        
        self.label_title = ctk.CTkLabel(self, text="Parametry vyhledávání", font=("Arial", 20, "bold"))
        self.label_title.pack(pady=10)

        self.entry_hledat = ctk.CTkEntry(self, placeholder_text="Co hledám (např. iphone 14 pro)", width=400)
        self.entry_hledat.pack(pady=5)

        self.entry_nechci = ctk.CTkEntry(self, placeholder_text="Co NECHCI (např. kryt, obal, poškozený) - oddělte čárkou", width=400)
        self.entry_nechci.pack(pady=5)

        frame_cena = ctk.CTkFrame(self, fg_color="transparent")
        frame_cena.pack(pady=5)
        self.entry_cena_od = ctk.CTkEntry(frame_cena, placeholder_text="Cena OD (Kč)", width=195)
        self.entry_cena_od.pack(side="left", padx=5)
        self.entry_cena_do = ctk.CTkEntry(frame_cena, placeholder_text="Cena DO (Kč)", width=195)
        self.entry_cena_do.pack(side="left", padx=5)

        frame_lokalita = ctk.CTkFrame(self, fg_color="transparent")
        frame_lokalita.pack(pady=5)
        self.entry_psc = ctk.CTkEntry(frame_lokalita, placeholder_text="PSČ středu", width=195)
        self.entry_psc.pack(side="left", padx=5)
        self.entry_okruh = ctk.CTkEntry(frame_lokalita, placeholder_text="Okruh (km)", width=195)
        self.entry_okruh.pack(side="left", padx=5)

        self.btn_hledat = ctk.CTkButton(self, text="Spustit vyhledávání", command=self.spustit_vlakno, font=("Arial", 16, "bold"), height=40)
        self.btn_hledat.pack(pady=20)

        self.log_box = ctk.CTkTextbox(self, width=480, height=300, state="disabled")
        self.log_box.pack(pady=10)

    def log(self, text):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", text + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def nacti_historii(self):
        if os.path.exists(self.soubor_historie):
            with open(self.soubor_historie, 'r', encoding='utf-8') as f:
                return set(line.strip() for line in f if line.strip())
        return set()

    def uloz_do_historie(self, nove_odkazy):
        if nove_odkazy:
            with open(self.soubor_historie, 'a', encoding='utf-8') as f:
                for odkaz in nove_odkazy:
                    f.write(odkaz + "\n")

    def spustit_vlakno(self):
        hledat = self.entry_hledat.get().strip()
        if not hledat:
            self.log("CHYBA: Musíte vyplnit pole 'Co hledám'!")
            return

        self.btn_hledat.configure(state="disabled", text="Stahuji data...")
        self.log_box.configure(state="normal")
        self.log_box.delete("0.0", "end")
        self.log_box.configure(state="disabled")
        
        thread = threading.Thread(target=self.scrape_bazos)
        thread.start()

    def scrape_bazos(self):
        hledat = self.entry_hledat.get().strip()
        nechci_text = self.entry_nechci.get().strip().lower()
        nechcene_razy = [slovo.strip() for slovo in nechci_text.split(',')] if nechci_text else []
        
        cena_od = self.entry_cena_od.get().strip()
        cena_do = self.entry_cena_do.get().strip()
        psc = self.entry_psc.get().strip()
        okruh = self.entry_okruh.get().strip()
        
        nazev_souboru = f"vysledky_PRO_{hledat.replace(' ', '_')}.xlsx"
        vyraz_url = hledat.replace(' ', '+')
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
        
        produkty = []
        nove_nalezene_odkazy = []
        historie = self.nacti_historii()
        limit_stranek = 10
        
        self.log(f"--- Zahajuji stahování dat z celého Bazoše ---")
        self.log(f"Načteno {len(historie)} inzerátů z historie paměti.")

        for strana in range(limit_stranek):
            offset = strana * 20
            strana_param = f"&strana={offset}" if offset > 0 else ""
            url = f"https://www.bazos.cz/search.php?hledat={vyraz_url}&rubriky=www&hlokalita={psc}&humkreis={okruh}&cenaod={cena_od}&cenado={cena_do}&Submit=Hledat{strana_param}"
            
            try:
                odpoved = requests.get(url, headers=headers)
                odpoved.raise_for_status()
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 404:
                    self.log(f"  Dosažen konec výsledků.")
                    break
                else:
                    self.log(f"  Chyba serveru: {e.response.status_code} - {e}")
                    break
            except Exception as e:
                self.log(f"  Chyba při načítání: {e}")
                break
            
            soup = BeautifulSoup(odpoved.text, 'html.parser')
            inzeraty = soup.find_all('div', class_='inzeraty')
            
            if not inzeraty:
                self.log(f"  Žádné další inzeráty na straně {strana + 1}.")
                break
                
            self.log(f"  Zpracovávám stranu {strana + 1}...")
            
            for inzerat in inzeraty:
                nadpis_h2 = inzerat.find('h2', class_='nadpis')
                if nadpis_h2:
                    odkaz_tag = nadpis_h2.find('a')
                    if odkaz_tag:
                        nazev = odkaz_tag.text.strip()
                        
                        preskocit = any(slovo in nazev.lower() for slovo in nechcene_razy)
                        if preskocit:
                            continue
                        
                        odkaz_href = odkaz_tag['href']
                        cisty_odkaz = odkaz_href if odkaz_href.startswith('http') else "https://www.bazos.cz" + odkaz_href
                        
                        je_novy = cisty_odkaz not in historie
                        if je_novy:
                            nove_nalezene_odkazy.append(cisty_odkaz)
                        
                        cena_div = inzerat.find('div', class_='inzeratycena')
                        cena = cena_div.text.strip() if cena_div else "Neznámá"
                        
                        lokalita_div = inzerat.find('div', class_='inzeratylok')
                        lokalita = lokalita_div.get_text(separator=', ', strip=True) if lokalita_div else "Neznámá"
                        
                        datum = "Neznámé"
                        zhlednuti = "Neznámé"
                        dalsi_info = inzerat.find_all('span', class_='velikost10')
                        for info in dalsi_info:
                            text_info = info.text.strip()
                            if text_info.startswith('[') and text_info.endswith(']'):
                                datum = text_info.strip('[]')
                            elif ' x' in text_info:
                                zhlednuti = text_info.replace(' x', '')
                        
                        produkty.append({
                            "Nový?": "ZCELA NOVÝ!" if je_novy else "-",
                            "Název inzerátu": nazev,
                            "Cena": cena,
                            "Lokalita": lokalita,
                            "Datum přidání": datum,
                            "Zhlédnutí": zhlednuti,
                            "Odkaz_skryty": cisty_odkaz
                        })
            
            time.sleep(1.5)

        self.ulozit_data(produkty, nazev_souboru)
        self.uloz_do_historie(nove_nalezene_odkazy)
        
        self.btn_hledat.configure(state="normal", text="Spustit vyhledávání")

    def ulozit_data(self, produkty, nazev_souboru):
        if not produkty:
            self.log("\nNebyly nalezeny žádné inzeráty odpovídající zadání.")
            return

        df = pd.DataFrame(produkty)
        df['Hodnota_ceny'] = df['Cena'].apply(cista_cena)
        df = df.sort_values(by='Hodnota_ceny').drop(columns=['Hodnota_ceny'])
        df['Odkaz'] = '=HYPERLINK("' + df['Odkaz_skryty'] + '", "' + df['Odkaz_skryty'] + '")'
        df = df.drop(columns=['Odkaz_skryty'])
        
        try:
            with pd.ExcelWriter(nazev_souboru, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Inzeráty')
                worksheet = writer.sheets['Inzeráty']
                
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 60
                worksheet.column_dimensions['C'].width = 15
                worksheet.column_dimensions['D'].width = 20
                worksheet.column_dimensions['E'].width = 15
                worksheet.column_dimensions['F'].width = 12
                worksheet.column_dimensions['G'].width = 60

                hlavicka_pozadi = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                hlavicka_pismo = Font(color="FFFFFF", bold=True)
                
                barva_suda = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
                barva_licha = PatternFill(start_color="D0E0FD", end_color="D0E0FD", fill_type="solid")
                barva_nova = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                tenka_cara = Side(border_style="thin", color="000000")
                ohraniceni = Border(left=tenka_cara, right=tenka_cara, top=tenka_cara, bottom=tenka_cara)
                zarovnani_stred = Alignment(horizontal="center", vertical="center")
                zarovnani_vlevo = Alignment(horizontal="left", vertical="center")

                for col_num in range(1, worksheet.max_column + 1):
                    bunka = worksheet.cell(row=1, column=col_num)
                    bunka.fill = hlavicka_pozadi
                    bunka.font = hlavicka_pismo
                    bunka.alignment = zarovnani_stred
                    bunka.border = ohraniceni

                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    je_zcela_novy = (row[0].value == "ZCELA NOVÝ!")
                    
                    for bunka in row:
                        bunka.border = ohraniceni
                        bunka.fill = barva_nova if je_zcela_novy else (barva_suda if bunka.row % 2 == 0 else barva_licha)
                        bunka.alignment = zarovnani_stred if bunka.column in [1, 3, 4, 5, 6] else zarovnani_vlevo
                            
                for row_num in range(1, worksheet.max_row + 1):
                    worksheet.row_dimensions[row_num].height = 22

                worksheet.freeze_panes = 'A2'
                worksheet.auto_filter.ref = worksheet.dimensions

            self.log(f"\n--- HOTOVO ---")
            self.log(f"Uloženo {len(df)} inzerátů do: {nazev_souboru}")
        except Exception as e:
            self.log(f"\nCHYBA PŘI UKLÁDÁNÍ: {e}")
            self.log("Zavřete Excel soubor, pokud ho máte právě otevřený!")

if __name__ == "__main__":
    app = BazosScraperApp()
    app.mainloop()